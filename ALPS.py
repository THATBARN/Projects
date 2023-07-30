import requests
from datetime import datetime
from dotenv import dotenv_values
import os

from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")

import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


env_conf = dotenv_values(os.path.abspath(os.path.dirname(__file__)) + "/.env")


# # GET new order ids
url = "https://admin.bozoraka.com/api/new-market-request-ids?sellerId=25"
headers = {
    'Content-type': 'application/json',
    'Authorization': env_conf.get("BOZORAKA_AUTH")
}
response = requests.get(url, headers=headers)
order_ids = response.text

# # POST order ids and receive order excel file
url = "https://admin.bozoraka.com/api/new-market-requests-excel"
response = requests.post(url, data=order_ids, headers=headers)

temp_file = env_conf.get("ORDER_FOLDER_PATH") + "/temp.xlsx"
with open(temp_file, 'wb') as e:
    e.write(response.content) # create temporary file of order excel spreadsheet

# # Edit excel spreadsheet to remove duplicates...
wb = load_workbook(filename="/Users/salmondehkanov/Desktop/Github/Orders/test.xlsx")
sheet = wb.active

os.remove(env_conf.get("ORDER_FOLDER_PATH") + "/temp.xlsx") # delete temp file, as it is now unneeded

# # Put all order numbers in list
column = sheet["A"]
order_numbers = []
for order in column:
    order = order.value
    order_numbers.append(order)
order_numbers.remove(None)

# # Cycle through list to find dupes

dupe_orders = []
for num in order_numbers:
    if num.count("-") > 2:
        dupe_orders.append(num)
        if num[0:-2] in order_numbers:
            if num[0:-2] not in dupe_orders:
                dupe_orders.append(num[0:-2])
print(dupe_orders)

# # Delete dupe orders
row = 1
while row <= sheet.max_row:
    if sheet.cell(row=row, column=1).value in dupe_orders:
        sheet.delete_rows(row, 1)
    else:
        row += 1


# # Set time and save file
now = datetime.now()
dt_string = "/" + now.strftime("%Y%m%d-%H%M") + ".xlsx" # datetime object containing current date and time

wb.save(env_conf.get("ORDER_FOLDER_PATH") + dt_string)


# # Upload file to ALPS
options = Options()
options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
options.headless = True

driver = webdriver.Chrome(env_conf.get("CHROME_DRIVER_PATH"), options=options)

driver.get(env_conf.get("ALPS_URL"))
driver.find_element(
    By.XPATH, '//*[@id="principal"]/input').send_keys(env_conf.get("ALPS_ID"))
driver.find_element(
    By.XPATH, '//*[@id="credential"]/input').send_keys(env_conf.get("ALPS_PASS"))

login = driver.find_element(By.XPATH, '//*[@id="btn-login"]')
login.click()

time.sleep(3)

button = driver.find_element(
    By.XPATH, '/html/body/div[3]/header/nav/div[1]/div[1]/i-button')
button.click()

button = driver.find_element(
    By.XPATH, '/html/body/div[3]/header/nav/div[3]/div/div[2]/div[7]/div[2]/ul/li/a')
button.click()

time.sleep(3)

workframe = driver.find_element(By.ID, "workframe_10782")
driver.switch_to.frame(workframe)
driver.execute_script("arguments[0].removeAttribute('readonly')",
                      WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="cboUsrFmat01"]/input[2]'))))

typeinput = driver.find_element(By.XPATH, '//*[@id="cboUsrFmat01"]/input[2]')
typeinput.send_keys(Keys.ARROW_DOWN, Keys.ARROW_DOWN, Keys.ARROW_DOWN, Keys.ENTER)

file_upload = driver.find_element(
    By.XPATH, '//*[@id="files"]'
)

file_upload.send_keys(env_conf.get("ORDER_FOLDER_PATH") + dt_string)

time.sleep(5)
